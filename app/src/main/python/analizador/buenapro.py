from flask import Blueprint, render_template, request, jsonify, send_file
import pdfplumber
import re
import os
import io
from datetime import datetime
import requests
import urllib3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from werkzeug.utils import secure_filename

from . import digesa, utiles

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

bp = Blueprint('buenapro', __name__, url_prefix='/buenapro')
import tempfile
UPLOAD_FOLDER = os.path.join(os.environ.get('HOME', tempfile.gettempdir()), 'uploads')

OSCE_BASE  = 'https://eap.osce.gob.pe/perfilprov-bus/1.0/ficha'

ALLOWED_EXTENSIONS = {'pdf'}

def allowed_file(f):
    return '.' in f and f.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def n(v):
    return ' '.join((v or '').split()).strip()


# ─────────────────────────────
#  EXTRACCIÓN PDF SEACE (Buena Pro)
# ─────────────────────────────

def extraer_datos_buena_pro(pdf_path, nombre_archivo):
    datos = {
        'archivo': nombre_archivo,
        'entidad': '',
        'nomenclatura': '',
        'descripcion': '',
        'ganadores': [],
        'error': ''
    }
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto = ''
            for pagina in pdf.pages:
                t = pagina.extract_text() or ''
                texto += t + '\n'

            m = re.search(r'Entidad\s+convocante\s*[:\-]?\s*(.+)', texto, re.IGNORECASE)
            if m:
                datos['entidad'] = n(m.group(1))

            m = re.search(r'Nomenclatura\s*[:\-]?\s*(\S+)', texto, re.IGNORECASE)
            if m:
                datos['nomenclatura'] = n(m.group(1))

            m = re.search(r'Descripci[oó]n\s+del\s+objeto\s*[:\-]?\s*(.+?)(?:\n(?:Nro\.?\s+Item|Entidad|Nomenclatura)|\Z)', texto, re.IGNORECASE | re.DOTALL)
            if m:
                datos['descripcion'] = n(m.group(1))[:300]

            # Buscar ganadores: RUC 11 dígitos + razón social
            patrones = [
                r'(\d{11})-([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑA-Za-z0-9&\-\.,\s]+(?:S\.A\.C\.|S\.A\.|E\.I\.R\.L\.|S\.R\.L\.|SAC|EIRL|SRL|LTDA|S\.R\.L|S\.A\.C|S\.A)\.?)',
                r'(\d{11})\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑA-Za-z0-9&\-\.,\s]+(?:S\.A\.C\.|S\.A\.|E\.I\.R\.L\.|S\.R\.L\.|SAC|EIRL|SRL|LTDA)\.?)',
                r'^(\d{11})\s+(.+?)(?:\s+[\d,\.]+\s*$)',
            ]

            vistos = set()
            for patron in patrones:
                for match in re.finditer(patron, texto, re.IGNORECASE | re.MULTILINE):
                    ruc = match.group(1)
                    razon = n(match.group(2))
                    if len(ruc) == 11 and ruc not in vistos and len(razon) > 4:
                        vistos.add(ruc)
                        pos = match.end()
                        frag = texto[pos:pos+150]
                        monto = ''
                        m2 = re.search(r'([\d,\.]{4,})', frag)
                        if m2:
                            monto = m2.group(1).replace(',', '')
                        datos['ganadores'].append({'ruc': ruc, 'razon_social': razon, 'monto': monto})

            if not datos['ganadores']:
                for ruc in re.findall(r'\b(\d{11})\b', texto):
                    if ruc not in vistos:
                        vistos.add(ruc)
                        datos['ganadores'].append({'ruc': ruc, 'razon_social': '', 'monto': ''})

    except Exception as e:
        datos['error'] = str(e)

    return datos


# ─────────────────────────────
#  OSCE
# ─────────────────────────────

def consultar_osce(ruc):
    ruc = re.sub(r'\D', '', ruc or '')
    vacio = {'ruc': ruc, 'razon_social': '', 'emails': [], 'telefonos': [],
             'celulares': [], 'estado_rnp': '', 'es_apto': False}
    if len(ruc) != 11:
        return vacio
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122 Safari/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'es-PE,es;q=0.9',
            'Referer': 'https://apps.osce.gob.pe/perfilprov-ui/buscar',
            'Origin': 'https://apps.osce.gob.pe',
        }
        r = requests.get(f'{OSCE_BASE}/{ruc}', headers=headers, timeout=15, verify=False)
        if r.status_code != 200:
            return vacio
        data = r.json()
        prov = data.get('proveedorT01', {})
        tels = prov.get('telefonos') or []
        emails = prov.get('emails') or []
        cel  = [t for t in tels if t.startswith('9') and len(re.sub(r'\D','',t)) == 9]
        fijo = [t for t in tels if t not in cel]
        return {
            'ruc': prov.get('numRuc', ruc),
            'razon_social': n(prov.get('nomRzsProv', '')),
            'emails': [e.lower() for e in emails],
            'telefonos': fijo,
            'celulares': cel,
            'estado_rnp': 'Habilitado' if prov.get('esHabilitado') else 'No habilitado',
            'es_apto': prov.get('esAptoContratar', False),
        }
    except Exception as e:
        print(f'[OSCE ERROR] {ruc}: {e}')
        return vacio


# ─────────────────────────────
#  DIGESA — la implementacion vive en digesa.py (compartida con contratos)
# ─────────────────────────────

def consultar_digesa(ruc):
    return digesa.consultar(ruc)




# ─────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────

def generar_excel(resultados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ganadores Buena Pro'

    color_header = PatternFill('solid', start_color='1a3a5c', end_color='1a3a5c')
    color_fab    = PatternFill('solid', start_color='d4edda', end_color='d4edda')
    color_rev    = PatternFill('solid', start_color='f8d7da', end_color='f8d7da')
    color_alt    = PatternFill('solid', start_color='f0f4f8', end_color='f0f4f8')
    font_h  = Font(color='FFFFFF', bold=True, size=10, name='Arial')
    font_b  = Font(size=10, name='Arial')
    ac = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    borde = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )

    encabezados = [
        'N°', 'Archivo PDF', 'Entidad Convocante', 'Nomenclatura', 'Descripción Objeto',
        'RUC Ganador', 'Razón Social',  'Monto Adj. (S/)',
        # OSCE
        'Estado RNP', 'Apto Contratar', 'Tel. OSCE', 'Celular OSCE', 'Email OSCE',
        # DIGESA
        'Tipo', 'Tiene DIGESA', 'N° Registros',
        'Registros DIGESA (Código | Producto | Vence)',
        'Tel. DIGESA', 'Rep. Legal DIGESA', 'Dirección DIGESA',
    ]

    for col, titulo in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.fill = color_header
        c.font = font_h
        c.alignment = ac
        c.border = borde
    ws.row_dimensions[1].height = 35

    fila = 2
    n_global = 0
    for res in resultados:
        ganadores = res.get('ganadores_detalle', [])
        if not ganadores:
            ganadores = [{'ruc': '', 'razon_social': '', 'monto': '', 'osce': {}, 'digesa': {}}]

        for g in ganadores:
            n_global += 1
            osce   = g.get('osce', {})
            digesa = g.get('digesa', {})
            tiene  = digesa.get('tiene_digesa', False)
            tipo   = digesa.get('tipo', '')

            reg_texto = '\n'.join(
                f"{r['numero']} | {r['producto'][:60]} | Vence: {r.get('fecha_vencimiento','')}"
                for r in digesa.get('registros', [])
            )

            valores = [
                n_global,
                res.get('archivo', ''),
                res.get('entidad', ''),
                res.get('nomenclatura', ''),
                res.get('descripcion', ''),
                g.get('ruc', ''),
                g.get('razon_social', '') or osce.get('razon_social', ''),
                g.get('monto', ''),
                osce.get('estado_rnp', ''),
                'Sí' if osce.get('es_apto') else 'No',
                ', '.join(osce.get('telefonos', [])),
                ', '.join(osce.get('celulares', [])),
                ', '.join(osce.get('emails', [])),
                tipo,
                'Sí' if tiene else 'No',
                len(digesa.get('registros', [])),
                reg_texto,
                ', '.join(digesa.get('telefonos', [])),
                digesa.get('rep_legal', ''),
                digesa.get('direccion', ''),
            ]

            fill_fila = color_fab if tiene else color_rev
            if fila % 2 == 0 and not tiene and tipo != 'Revendedor':
                fill_fila = color_alt

            for col, val in enumerate(valores, 1):
                c = ws.cell(row=fila, column=col, value=val)
                c.font = font_b
                c.border = borde
                c.alignment = ac if col in (1, 10, 15, 16) else al
                c.fill = fill_fila

            ws.row_dimensions[fila].height = max(30, min(120, len(reg_texto or '') // 3))
            fila += 1

    anchos = [5, 25, 35, 20, 40, 14, 45, 16, 16, 10, 14, 14, 28, 22, 10, 10, 60, 14, 35, 40]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = ancho

    ws.freeze_panes = 'A2'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _safe_pdf_text(value):
    txt = n(value)
    txt = txt.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    return txt


def generar_pdf_general(resultados):
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    style_title = styles['Heading1']
    style_sub = styles['Heading3']
    style_normal = styles['BodyText']
    style_small = styles['BodyText'].clone('small_pdf')
    style_small.fontSize = 8
    style_small.leading = 10
    style_tiny = styles['BodyText'].clone('tiny_pdf')
    style_tiny.fontSize = 7
    style_tiny.leading = 9

    rows = []
    n_global = 0
    fab = 0
    rev = 0
    rnp_hab = 0

    for res in resultados:
        ganadores = res.get('ganadores_detalle', [])
        for g in ganadores:
            n_global += 1
            osce = g.get('osce', {})
            digesa = g.get('digesa', {})

            tipo = digesa.get('tipo', '') or 'Revendedor'
            if digesa.get('tiene_digesa'):
                fab += 1
            else:
                rev += 1

            estado_rnp = osce.get('estado_rnp', '')
            if estado_rnp == 'Habilitado':
                rnp_hab += 1

            tels = ', '.join((osce.get('telefonos') or []) + (osce.get('celulares') or []))
            emails = ', '.join(osce.get('emails') or [])
            contacto = tels or ', '.join(digesa.get('telefonos') or []) or '-'

            rows.append([
                str(n_global),
                _safe_pdf_text(res.get('archivo', '')),
                _safe_pdf_text(g.get('ruc', '')),
                _safe_pdf_text(g.get('razon_social') or osce.get('razon_social', '')),
                _safe_pdf_text(g.get('monto', '')),
                _safe_pdf_text(tipo),
                str(len(digesa.get('registros', []) or [])),
                _safe_pdf_text(estado_rnp or '-'),
                _safe_pdf_text(contacto),
                _safe_pdf_text(emails or '-'),
            ])

    elementos = []
    elementos.append(Paragraph('Reporte General de Buena Pro', style_title))
    elementos.append(
        Paragraph(
            f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
            style_normal,
        )
    )
    elementos.append(Spacer(1, 6))
    elementos.append(Paragraph('Resumen', style_sub))
    elementos.append(Paragraph(f'PDFs procesados: {len(resultados)}', style_normal))
    elementos.append(Paragraph(f'Ganadores encontrados: {n_global}', style_normal))
    elementos.append(Paragraph(f'Fabricantes / Distribuidores: {fab}', style_normal))
    elementos.append(Paragraph(f'Revendedores: {rev}', style_normal))
    elementos.append(Paragraph(f'Habilitados en RNP (OSCE): {rnp_hab}', style_normal))
    elementos.append(Spacer(1, 10))

    if rows:
        header = ['N°', 'Archivo', 'RUC', 'Razón Social', 'Monto', 'Tipo', 'Reg. DIGESA', 'RNP', 'Contacto', 'Email']
        table_data = [header]
        for row in rows:
            table_data.append([
                Paragraph(row[0], style_small),
                Paragraph(row[1], style_small),
                Paragraph(row[2], style_small),
                Paragraph(row[3], style_small),
                Paragraph(row[4], style_small),
                Paragraph(row[5], style_small),
                Paragraph(row[6], style_small),
                Paragraph(row[7], style_small),
                Paragraph(row[8], style_small),
                Paragraph(row[9], style_small),
            ])

        tabla = Table(
            table_data,
            colWidths=[10 * mm, 30 * mm, 20 * mm, 47 * mm, 18 * mm, 28 * mm, 18 * mm, 18 * mm, 42 * mm, 42 * mm],
            repeatRows=1,
        )
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2a44')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
            ('ALIGN', (6, 0), (7, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#c7d2fe')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#eef2ff')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elementos.append(tabla)
        elementos.append(Spacer(1, 10))

        elementos.append(Paragraph('Detalle Completo por Ganador', style_sub))
        elementos.append(Spacer(1, 4))

        item = 0
        for res in resultados:
            ganadores = res.get('ganadores_detalle', [])
            for g in ganadores:
                item += 1
                osce = g.get('osce', {})
                digesa = g.get('digesa', {})
                registros = digesa.get('registros', []) or []

                elementos.append(Paragraph(
                    f"{item}. {_safe_pdf_text(g.get('razon_social') or osce.get('razon_social', 'SIN RAZON SOCIAL'))}",
                    style_sub
                ))

                detalle_data = [
                    [Paragraph('<b>Archivo</b>', style_small), Paragraph(_safe_pdf_text(res.get('archivo', '-')), style_small),
                     Paragraph('<b>Entidad</b>', style_small), Paragraph(_safe_pdf_text(res.get('entidad', '-')), style_small)],
                    [Paragraph('<b>Nomenclatura</b>', style_small), Paragraph(_safe_pdf_text(res.get('nomenclatura', '-')), style_small),
                     Paragraph('<b>Descripcion</b>', style_small), Paragraph(_safe_pdf_text(res.get('descripcion', '-')), style_small)],
                    [Paragraph('<b>RUC</b>', style_small), Paragraph(_safe_pdf_text(g.get('ruc', '-')), style_small),
                     Paragraph('<b>Monto</b>', style_small), Paragraph(_safe_pdf_text(g.get('monto', '-')), style_small)],
                    [Paragraph('<b>Estado RNP</b>', style_small), Paragraph(_safe_pdf_text(osce.get('estado_rnp', '-')), style_small),
                     Paragraph('<b>Apto Contratar</b>', style_small), Paragraph('Si' if osce.get('es_apto') else 'No', style_small)],
                    [Paragraph('<b>Telefonos OSCE</b>', style_small), Paragraph(_safe_pdf_text(', '.join(osce.get('telefonos', [])) or '-'), style_small),
                     Paragraph('<b>Celulares OSCE</b>', style_small), Paragraph(_safe_pdf_text(', '.join(osce.get('celulares', [])) or '-'), style_small)],
                    [Paragraph('<b>Emails OSCE</b>', style_small), Paragraph(_safe_pdf_text(', '.join(osce.get('emails', [])) or '-'), style_small),
                     Paragraph('<b>Tipo DIGESA</b>', style_small), Paragraph(_safe_pdf_text(digesa.get('tipo', '-')), style_small)],
                    [Paragraph('<b>Tiene DIGESA</b>', style_small), Paragraph('Si' if digesa.get('tiene_digesa') else 'No', style_small),
                     Paragraph('<b>Registros DIGESA</b>', style_small), Paragraph(str(len(registros)), style_small)],
                    [Paragraph('<b>Telefonos DIGESA</b>', style_small), Paragraph(_safe_pdf_text(', '.join(digesa.get('telefonos', [])) or '-'), style_small),
                     Paragraph('<b>Rep. Legal DIGESA</b>', style_small), Paragraph(_safe_pdf_text(digesa.get('rep_legal', '-')), style_small)],
                    [Paragraph('<b>Direccion DIGESA</b>', style_small), Paragraph(_safe_pdf_text(digesa.get('direccion', '-')), style_small),
                     Paragraph('<b></b>', style_small), Paragraph('', style_small)],
                ]

                tabla_detalle = Table(detalle_data, colWidths=[30 * mm, 58 * mm, 34 * mm, 125 * mm])
                tabla_detalle.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#eef2ff')),
                    ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#eef2ff')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]))
                elementos.append(tabla_detalle)
                elementos.append(Spacer(1, 5))

                elementos.append(Paragraph('Registros DIGESA (numero, producto y toda la informacion)', style_small))
                if registros:
                    reg_table_data = [[
                        Paragraph('<b>N</b>', style_tiny),
                        Paragraph('<b>Numero</b>', style_tiny),
                        Paragraph('<b>Producto</b>', style_tiny),
                        Paragraph('<b>Producto Completo</b>', style_tiny),
                        Paragraph('<b>F. Emision</b>', style_tiny),
                        Paragraph('<b>F. Vencimiento</b>', style_tiny),
                        Paragraph('<b>Empresa</b>', style_tiny),
                        Paragraph('<b>Direccion Registro</b>', style_tiny),
                        Paragraph('<b>Anio</b>', style_tiny),
                    ]]

                    for i, reg in enumerate(registros, start=1):
                        reg_table_data.append([
                            Paragraph(str(i), style_tiny),
                            Paragraph(_safe_pdf_text(reg.get('numero', '-')), style_tiny),
                            Paragraph(_safe_pdf_text(reg.get('producto', '-')), style_tiny),
                            Paragraph(_safe_pdf_text(reg.get('producto_completo', '-')), style_tiny),
                            Paragraph(_safe_pdf_text(reg.get('fecha_emision', '-')), style_tiny),
                            Paragraph(_safe_pdf_text(reg.get('fecha_vencimiento', '-')), style_tiny),
                            Paragraph(_safe_pdf_text(reg.get('empresa', '-')), style_tiny),
                            Paragraph(_safe_pdf_text(reg.get('direccion_registro', '-')), style_tiny),
                            Paragraph(_safe_pdf_text(reg.get('anio', '-')), style_tiny),
                        ])

                    tabla_regs = Table(
                        reg_table_data,
                        colWidths=[8 * mm, 22 * mm, 36 * mm, 58 * mm, 18 * mm, 22 * mm, 34 * mm, 54 * mm, 10 * mm],
                        repeatRows=1,
                    )
                    tabla_regs.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2a44')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('ALIGN', (0, 0), (1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                        ('LEFTPADDING', (0, 0), (-1, -1), 3),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ]))
                    elementos.append(tabla_regs)
                else:
                    elementos.append(Paragraph('Sin registros DIGESA para este RUC.', style_small))

                elementos.append(Spacer(1, 10))
    else:
        elementos.append(Paragraph('No hay datos para exportar.', style_normal))

    doc.build(elementos)
    output.seek(0)
    return output


# ─────────────────────────────
#  RUTAS FLASK
# ─────────────────────────────

@bp.route('/')
def index():
    return render_template('buenapro.html')

@bp.route('/procesar', methods=['POST'])
def procesar():
    if 'pdfs' not in request.files:
        return jsonify({'error': 'No se enviaron archivos'}), 400
    archivos = request.files.getlist('pdfs')
    resultados = []
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    for archivo in archivos:
        if archivo and allowed_file(archivo.filename):
            nombre = secure_filename(archivo.filename)
            ruta = os.path.join(UPLOAD_FOLDER, nombre)
            archivo.save(ruta)
            datos = extraer_datos_buena_pro(ruta, nombre)
            resultados.append(datos)
            os.remove(ruta)
    return jsonify({'resultados': resultados})

@bp.route('/consultar_ganador', methods=['POST'])
def consultar_ganador():
    body = request.json or {}
    ruc  = (body.get('ruc') or '').strip()
    if not re.fullmatch(r'\d{11}', ruc):
        return jsonify({'error': 'RUC inválido'}), 400
    osce   = consultar_osce(ruc)
    digesa = consultar_digesa(ruc)
    return jsonify({'osce': osce, 'digesa': digesa})

@bp.route('/exportar', methods=['POST'])
def exportar():
    body = request.json or {}
    datos = body.get('datos', [])
    excel = generar_excel(datos)
    return send_file(
        excel,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=utiles.nombre_archivo('analisis_buena_pro', 'xlsx')
    )


@bp.route('/exportar_pdf', methods=['POST'])
def exportar_pdf():
    body = request.json or {}
    datos = body.get('datos', [])
    pdf = generar_pdf_general(datos)
    return send_file(
        pdf,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=utiles.nombre_archivo('analisis_buena_pro_general', 'pdf')
    )


def procesar_archivo(ruta, nombre):
    datos = extraer_datos_buena_pro(ruta, nombre)
    detalle = []
    for g in datos.get('ganadores', []):
        item = dict(g)
        item['osce'] = consultar_osce(g['ruc'])
        item['digesa'] = consultar_digesa(g['ruc'])
        detalle.append(item)
    datos['ganadores_detalle'] = detalle
    return datos
