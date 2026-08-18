from flask import Blueprint, render_template, request, jsonify, send_file
import pdfplumber
import re
import os
import io
import shutil
import sys
import time
import openpyxl
import requests
import urllib3
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.utils import secure_filename
from fpdf import FPDF

from . import digesa as digesa_api, utiles

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

try:
    import pytesseract
except ImportError:
    pytesseract = None

if load_dotenv:
    load_dotenv()

bp = Blueprint('contratos', __name__, url_prefix='/contratos')
import tempfile
UPLOAD_FOLDER = os.path.join(os.environ.get('HOME', tempfile.gettempdir()), 'uploads')

BASE = 'https://eap.osce.gob.pe/perfilprov-bus/1.0/ficha'

HEADERS = {
    'User-Agent'     : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122 Safari/537.36',
    'Accept'         : 'application/json, text/plain, */*',
    'Accept-Language': 'es-PE,es;q=0.9',
    'Referer'        : 'https://apps.osce.gob.pe/perfilprov-ui/buscar',
    'Origin'         : 'https://apps.osce.gob.pe',
}

ALLOWED_EXTENSIONS = {'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def normalizar_texto(valor):
    return ' '.join((valor or '').split()).strip()

def extraer_telefonos_desde_texto(texto):
    patron = r'(?:\+?51\s*)?(?:\(?\d{2,3}\)?[\s-]*)?\d{3,4}[\s-]?\d{3,4}'
    telefonos = []
    vistos = set()
    for raw in re.findall(patron, texto):
        digitos = re.sub(r'\D', '', raw)
        if len(digitos) < 7:
            continue
        valor = normalizar_texto(raw)
        if digitos not in vistos:
            vistos.add(digitos)
            telefonos.append(valor)
    return telefonos




#  MÓDULO OSCE
# ─────────────────────────────────────────────
def _n(v):
    return ' '.join(str(v or '').split()).strip()

def consultar_osce_por_ruc(ruc: str) -> dict:
    ruc = re.sub(r'\D', '', ruc)
    vacio = {
        'ruc': ruc, 'razon_social': '', 'emails': [],
        'telefonos_osce': [], 'celulares_osce': [],
        'direccion': '', 'estado_rnp': '',
        'capacidad_contratacion': '', 'es_apto': False
    }
    if len(ruc) != 11:
        return vacio

    try:
        data = {}
        for intento in range(2):
            try:
                r = requests.get(
                    f'{BASE}/{ruc}',
                    headers=HEADERS,
                    timeout=15,
                    verify=False
                )
                if r.status_code != 200:
                    print(f'[OSCE] status {r.status_code} para {ruc}')
                    if intento == 0:
                        time.sleep(0.35)
                        continue
                    return vacio
                data = r.json()
                break
            except Exception:
                if intento == 0:
                    time.sleep(0.35)
                    continue
                raise

        prov = data.get('proveedorT01') or {}
        if not prov:
            return vacio

        telefonos = prov.get('telefonos') or []
        emails    = prov.get('emails') or []

        celulares = [t for t in telefonos if t.startswith('9') and len(re.sub(r'\D','',t)) == 9]
        fijos     = [t for t in telefonos if t not in celulares]

        return {
            'ruc'                   : prov.get('numRuc', ruc),
            'razon_social'          : _n(prov.get('nomRzsProv', '')),
            'emails'                : [e.lower() for e in emails],
            'telefonos_osce'        : fijos,
            'celulares_osce'        : celulares,
            'direccion'             : '',
            'estado_rnp'            : 'Habilitado' if prov.get('esHabilitado') else 'No habilitado',
            'es_apto'               : prov.get('esAptoContratar', False),
            'capacidad_contratacion': '',
        }
    except Exception as e:
        print(f'[OSCE ERROR] {ruc}: {e}')
        return vacio

# ─────────────────────────────────────────────
#  Funciones para Consorcio y PDF
# ─────────────────────────────────────────────

def extraer_rucs_consorcio(texto_miembros):
    """Extrae RUCs de los miembros del consorcio"""
    rucs = []
    if not texto_miembros:
        return rucs
    for linea in texto_miembros.split('\n'):
        linea = linea.strip()
        if linea:
            ruc_match = re.search(r'(\d{11})', linea)
            if ruc_match:
                rucs.append(ruc_match.group(1))
    return rucs

def consultar_all_osce(ruc_principal, miembros_consorcio_texto):
    """Consulta OSCE para el RUC principal y todos los miembros del consorcio"""
    resultados = []
    rucs_consultados = set()
    
    # Consultar RUC principal
    if ruc_principal:
        resultados.append(consultar_osce_por_ruc(ruc_principal))
        rucs_consultados.add(ruc_principal)
    
    # Consultar RUCs del consorcio
    rucs_consorcio = extraer_rucs_consorcio(miembros_consorcio_texto)
    for ruc in rucs_consorcio:
        if ruc not in rucs_consultados:
            resultados.append(consultar_osce_por_ruc(ruc))
            rucs_consultados.add(ruc)
    
    return resultados

def generar_pdf_reporte(datos_completos):
    """Genera un PDF con formato profesional de los datos extraidos"""
    from fpdf import FPDF, XPos, YPos
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 15, "REPORTE DE LICITACIÓN", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    # Información Principal
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 10, "INFORMACIÓN PRINCIPAL", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=1)
    
    pdf.set_font("Helvetica", "", 10)
    datos_principales = [
        ("Archivo", datos_completos.get('archivo', 'N/A')),
        ("Número de Contrato", datos_completos.get('numero_contrato', 'N/A')),
        ("Zona", datos_completos.get('zona', 'N/A')),
    ]
    
    for label, valor in datos_principales:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(50, 7, f"{label}:")
        pdf.set_font("Helvetica", "", 9)
        valor_texto = str(valor)[:100]
        pdf.multi_cell(0, 7, valor_texto, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    pdf.cell(0, 5, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    # Información del Ganador
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 10, "CONTRATISTA GANADOR", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=1)
    
    pdf.set_font("Helvetica", "", 10)
    ruc_principal = datos_completos.get('ruc', '')
    contratista = datos_completos.get('contratista', '')
    rep_legal = datos_completos.get('representante_legal', '')
    
    datos_ganador = [
        ("RUC", ruc_principal),
        ("Razón Social", contratista),
        ("Representante Legal", rep_legal),
        ("DNI Representante", datos_completos.get('dni_representante', 'N/A')),
    ]
    
    for label, valor in datos_ganador:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(50, 7, f"{label}:")
        pdf.set_font("Helvetica", "", 9)
        valor_texto = str(valor)[:100]
        pdf.multi_cell(0, 7, valor_texto, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    pdf.cell(0, 5, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    # Información OSCE del Ganador
    if ruc_principal:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, "INFORMACIÓN OSCE - GANADOR", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=1)
        
        pdf.set_font("Helvetica", "", 9)
        osce_info = [
            ("Razón Social OSCE", datos_completos.get('osce_razon_social', 'N/A')),
            ("Estado RNP", datos_completos.get('osce_estado_rnp', 'N/A')),
            ("Apto para contratar", "Sí" if datos_completos.get('osce_es_apto') else "No"),
            ("Teléfono", ', '.join(datos_completos.get('osce_telefonos', ['N/A']))),
            ("Celular", ', '.join(datos_completos.get('osce_celulares', ['N/A']))),
            ("Email", ', '.join(datos_completos.get('osce_emails', ['N/A']))),
        ]
        
        for label, valor in osce_info:
            pdf.set_font("Helvetica", "B", 8.5)
            pdf.cell(45, 6, f"{label}:")
            pdf.set_font("Helvetica", "", 8.5)
            valor_texto = str(valor)[:120]
            pdf.multi_cell(0, 6, valor_texto, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        pdf.cell(0, 3, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    # Miembros del Consorcio
    miembros = datos_completos.get('miembros_consorcio', '').strip()
    if miembros:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, "MIEMBROS DEL CONSORCIO", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=1)
        
        pdf.set_font("Helvetica", "", 9)
        miembros_list = miembros.split('\n')
        for i, miembro in enumerate(miembros_list, 1):
            pdf.cell(10, 6, f"{i}.")
            pdf.multi_cell(0, 6, miembro.strip(), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        # Agregar info OSCE de miembros si está disponible
        if 'osce_consorcio' in datos_completos and datos_completos['osce_consorcio']:
            pdf.cell(0, 3, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(0, 9, "INFORMACIÓN OSCE - MIEMBROS", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=1)
            
            for osce_miembro in datos_completos['osce_consorcio']:
                pdf.set_font("Helvetica", "B", 8)
                ruc_miembro = osce_miembro.get('ruc', 'N/A')
                razon_miembro = osce_miembro.get('razon_social', 'N/A')
                pdf.cell(0, 6, f"RUC: {ruc_miembro} - {razon_miembro}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                pdf.set_font("Helvetica", "", 7.5)
                pdf.cell(10, 5, "")
                pdf.cell(40, 5, "Estado RNP:")
                pdf.cell(0, 5, osce_miembro.get('estado_rnp', 'N/A'), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                pdf.cell(10, 5, "")
                pdf.cell(40, 5, "Apto:")
                pdf.cell(0, 5, "Sí" if osce_miembro.get('es_apto') else "No", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                telefonos = ', '.join(osce_miembro.get('telefonos_osce', ['N/A']))
                pdf.cell(10, 5, "")
                pdf.cell(40, 5, "Teléfono:")
                pdf.multi_cell(0, 5, telefonos, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                celulares = ', '.join(osce_miembro.get('celulares_osce', ['N/A']))
                pdf.cell(10, 5, "")
                pdf.cell(40, 5, "Celular:")
                pdf.multi_cell(0, 5, celulares, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                pdf.cell(0, 3, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        pdf.cell(0, 5, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return output

# ─────────────────────────────────────────────
#  OCR
# ─────────────────────────────────────────────

def configurar_ocr():
    if pytesseract is None:
        return False, f"OCR no disponible ({sys.executable}): instala pytesseract."
    tesseract_cmd = os.getenv('TESSERACT_CMD', '').strip()
    if tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
    else:
        ruta = shutil.which('tesseract')
        if ruta:
            pytesseract.pytesseract.tesseract_cmd = ruta
        else:
            for p in ['C:/Program Files/Tesseract-OCR/tesseract.exe', 'C:/Program Files (x86)/Tesseract-OCR/tesseract.exe']:
                if os.path.exists(p):
                    pytesseract.pytesseract.tesseract_cmd = p
                    break
    try:
        v = pytesseract.get_tesseract_version()
        return True, f"OCR habilitado: Tesseract {v}."
    except Exception:
        return False, "pytesseract instalado pero Tesseract no encontrado. Configura TESSERACT_CMD."

def extraer_texto_ocr_pagina(pagina):
    if pytesseract is None:
        return ''
    try:
        imagen = pagina.to_image(resolution=300).original
        return (pytesseract.image_to_string(imagen, config='--oem 1 --psm 6 -l spa+eng') or '').strip()
    except Exception:
        return ''


# ─────────────────────────────────────────────
#  Extracción PDF
# ─────────────────────────────────────────────

def _limpiar_nombre_empresa(nombre):
    n = ' '.join(nombre.split()).strip(' ,.;:-')
    n = re.sub(r'^(?:LA\s+EMPRESA|EMPRESA|EL\s+CONSORCIO|CONSORCIO)\s+', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s+(?:REPRESENTADA\s+POR|CON\s+DOMICILIO|IDENTIFICADA\s+CON).*$', '', n, flags=re.IGNORECASE)
    return n.strip(' ,.;:-')

def extraer_miembros_consorcio(texto):
    if not re.search(r'\bconsorcio\b', texto, re.IGNORECASE):
        return ''
    miembros = []
    vistos = set()
    patrones = [
        r'([A-ZÁÉÍÓÚÑ0-9&\-\.,\s]{4,}?)\s+(?:con\s+)?RUC\s*[N°º]*\s*[:\.]?\s*(\d{11})',
        r'RUC\s*[N°º]*\s*[:\.]?\s*(\d{11})\s*(?:,|;|-)?\s*(?:de\s+la\s+empresa\s+|de\s+)?([A-ZÁÉÍÓÚÑ0-9&\-\.,\s]{4,})'
    ]
    for patron in patrones:
        for match in re.findall(patron, texto, re.IGNORECASE):
            nombre, ruc = match if patron == patrones[0] else (match[1], match[0])
            nombre = _limpiar_nombre_empresa(nombre)
            if not nombre or 'CONTRATISTA' in nombre.upper() or 'CONSORCIO' in nombre.upper():
                continue
            clave = (ruc, nombre.upper())
            if clave not in vistos:
                vistos.add(clave)
                miembros.append(f'{ruc} - {nombre}')
    return '\n'.join(miembros)

def extraer_datos_contrato(pdf_path, nombre_archivo):
    datos = {
        'archivo': nombre_archivo, 'numero_contrato': '', 'zona': '',
        'ruc': '', 'contratista': '', 'miembros_consorcio': '',
        'digesa_registros': [], 'digesa_productos': [], 'digesa_telefonos': [],
        'osce_razon_social': '', 'osce_emails': [], 'osce_telefonos': [],
        'osce_celulares': [], 'osce_direccion': '', 'osce_estado_rnp': '',
        'osce_capacidad': '', 'osce_es_apto': False,
        'ocr_aplicado': False, 'representante_legal': '', 'dni_representante': '', 'texto_raw': ''
    }
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto = ''
            for pagina in pdf.pages[:2]:
                texto_pagina = (pagina.extract_text() or '').strip()
                if len(texto_pagina) < 20:
                    texto_ocr = extraer_texto_ocr_pagina(pagina)
                    if texto_ocr:
                        datos['ocr_aplicado'] = True
                        texto_pagina = f"{texto_pagina}\n{texto_ocr}" if texto_pagina else texto_ocr
                if texto_pagina:
                    texto += texto_pagina + '\n'
            datos['texto_raw'] = texto

            m = re.search(r'CONTRATO\s+N[°º]\s*([^\n]+)', texto, re.IGNORECASE)
            if m:
                datos['numero_contrato'] = m.group(1).strip()
                zm = re.search(r'CGC[-\s]+(.+?)(?:/|$)', datos['numero_contrato'], re.IGNORECASE)
                if zm: datos['zona'] = zm.group(1).strip()

            m = re.search(r'RUC\s*[N°º]*\s*[:\.]?\s*(\d{11})', texto, re.IGNORECASE)
            if m: datos['ruc'] = m.group(1)

            m = re.search(r'de la otra parte\s+(.+?)\s+con\s+(?:código de\s+)?CONTRATISTA', texto, re.IGNORECASE | re.DOTALL)
            if m: datos['contratista'] = ' '.join(m.group(1).split()).strip()

            if not datos['contratista']:
                matches = re.findall(r'([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑA-Za-z\s\.]+(?:S\.A\.C\.|S\.A\.|E\.I\.R\.L\.|S\.R\.L\.|SAC|EIRL|SRL))', texto)
                if matches: datos['contratista'] = matches[-1].strip()

            m = re.search(r'(?:representante legal|Apoderado Legal)[^\n]*?(?:EL\(LA\)\s*)?(?:SEÑOR\(A\)|SEÑOR|SEÑORA)?\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑA-Za-z\s]+?)(?:,\s*con\s*DNI|con\s*DNI)', texto, re.IGNORECASE)
            if m: datos['representante_legal'] = ' '.join(m.group(1).split()).strip()

            matches_dni = re.findall(r'DNI\s*[N°º]*\s*[:\.]?\s*(\d{8})', texto, re.IGNORECASE)
            if matches_dni: datos['dni_representante'] = matches_dni[-1]

            datos['miembros_consorcio'] = extraer_miembros_consorcio(texto)
    except Exception as e:
        datos['error'] = str(e)
    return datos


# ─────────────────────────────────────────────
#  Excel
# ─────────────────────────────────────────────

def generar_excel(lista_datos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contratos"

    encabezados = [
        'N°', 'Archivo PDF', 'Número de Contrato', 'Zona', 'RUC',
        'Contratista (Ganador)', 'Miembros del Consorcio (RUC - Nombre)',
        'Representante Legal', 'DNI Representante',
        'Registros DIGESA', 'Productos DIGESA', 'Teléfonos DIGESA',
        'Razón Social OSCE', 'Email OSCE', 'Teléfonos OSCE', 'Celulares OSCE',
        'Estado RNP', 'Apto para contratar'
    ]

    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font_header = Font(color="FFFFFF", bold=True, size=11)

    for col, titulo in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.fill = fill_header
        celda.font = font_header
        celda.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 20
    fill_par = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for i, datos in enumerate(lista_datos, 1):
        digesa_regs = datos.get('digesa_registros', [])
        digesa_texto = '\n'.join(
            f"{r.get('numero_registro','')} | {r.get('estado','')} | {r.get('fecha_emision','')}"
            for r in digesa_regs
        ) if digesa_regs else ''

        valores = [
            i,
            datos.get('archivo', ''),
            datos.get('numero_contrato', ''),
            datos.get('zona', ''),
            datos.get('ruc', ''),
            datos.get('contratista', ''),
            datos.get('miembros_consorcio', ''),
            datos.get('representante_legal', ''),
            datos.get('dni_representante', ''),
            digesa_texto,
            '\n'.join(datos.get('digesa_productos', [])),
            '\n'.join(datos.get('digesa_telefonos', [])),
            datos.get('osce_razon_social', ''),
            '\n'.join(datos.get('osce_emails', [])),
            '\n'.join(datos.get('osce_telefonos', [])),
            '\n'.join(datos.get('osce_celulares', [])),
            datos.get('osce_estado_rnp', ''),
            'Sí' if datos.get('osce_es_apto') else 'No',
        ]

        for col, val in enumerate(valores, 1):
            celda = ws.cell(row=i + 1, column=col, value=val)
            if i % 2 == 0:
                celda.fill = fill_par
            celda.alignment = Alignment(vertical='center', wrap_text=True)

    anchos = [5, 30, 50, 25, 15, 40, 55, 35, 15, 35, 40, 20, 40, 30, 20, 20, 20, 15]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = ancho

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output



def procesar_archivo(ruta, nombre):
    datos = extraer_datos_contrato(ruta, nombre)
    ruc = (datos.get('ruc') or '').strip()
    if re.fullmatch(r'\d{11}', ruc):
        osce = consultar_osce_por_ruc(ruc)
        datos['osce_razon_social'] = osce.get('razon_social', '')
        datos['osce_emails'] = osce.get('emails', [])
        datos['osce_telefonos'] = osce.get('telefonos_osce', [])
        datos['osce_celulares'] = osce.get('celulares_osce', [])
        datos['osce_direccion'] = osce.get('direccion', '')
        datos['osce_estado_rnp'] = osce.get('estado_rnp', '')
        datos['osce_capacidad'] = osce.get('capacidad_contratacion', '')
        datos['osce_es_apto'] = osce.get('es_apto', False)
        digesa = consultar_digesa(ruc)
        datos['digesa_registros'] = digesa.get('registros', [])
        datos['digesa_telefonos'] = digesa.get('telefonos', [])
        datos['digesa_rep_legal'] = digesa.get('rep_legal', '')
        datos['digesa_direccion'] = digesa.get('direccion', '')
        datos['digesa_tiene'] = digesa.get('tiene_digesa', False)
        datos['digesa_tipo'] = digesa.get('tipo', 'Revendedor')
        datos['digesa_productos'] = list(dict.fromkeys(
            r.get('producto', '') for r in digesa.get('registros', []) if r.get('producto', '')
        ))
        miembros_texto = datos.get('miembros_consorcio', '')
        if miembros_texto:
            datos['osce_consorcio'] = consultar_all_osce(ruc, miembros_texto)[1:]
        else:
            datos['osce_consorcio'] = []
    return datos


# ─────────────────────────────────────────────
#  Rutas Flask
# ─────────────────────────────────────────────

@bp.route('/')
def index():
    return render_template('contratos.html')

@bp.route('/procesar', methods=['POST'])
def procesar():
    if 'pdfs' not in request.files:
        return jsonify({'error': 'No se enviaron archivos'}), 400
    archivos = request.files.getlist('pdfs')
    resultados = []
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    for archivo in archivos:
        if archivo and allowed_file(archivo.filename):
            nombre = secure_filename(archivo.filename)
            ruta = os.path.join(UPLOAD_FOLDER, nombre)
            archivo.save(ruta)
            resultados.append(procesar_archivo(ruta, nombre))
            os.remove(ruta)
    return jsonify({'resultados': resultados})

@bp.route('/exportar', methods=['POST'])
def exportar():
    datos = request.json.get('datos', [])
    excel_file = generar_excel(datos)
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=utiles.nombre_archivo('contratos_ganadores', 'xlsx')
    )

@bp.route('/exportar_pdf', methods=['POST'])
def exportar_pdf():
    """Exporta un registro individual a PDF"""
    datos = request.json.get('datos', {})
    if not datos:
        return jsonify({'error': 'No hay datos para exportar'}), 400
    
    try:
        pdf_file = generar_pdf_reporte(datos)
        archivo_nombre = utiles.nombre_archivo(f"reporte_{datos.get('ruc', 'sin_ruc')}", 'pdf')
        return send_file(
            pdf_file,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=archivo_nombre
        )
    except Exception as e:
        return jsonify({'error': f'Error generando PDF: {str(e)}'}), 500

@bp.route('/exportar_pdf_lote', methods=['POST'])
def exportar_pdf_lote():
    """Exporta múltiples registros en un PDF consolidado con toda la información"""
    from fpdf import FPDF, XPos, YPos
    
    datos_lista = request.json.get('datos', [])
    if not datos_lista:
        return jsonify({'error': 'No hay datos para exportar'}), 400
    
    try:
        pdf = FPDF()
        
        for idx, datos in enumerate(datos_lista):
            if idx > 0:
                pdf.add_page()
            else:
                pdf.add_page()
            
            # ENCABEZADO
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 12, f"REPORTE DE LICITACIÓN #{idx + 1}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_line_width(0.5)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.cell(0, 2, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            
            # INFORMACIÓN PRINCIPAL
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(0, 8, "INFORMACIÓN PRINCIPAL DEL CONTRATO", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=0)
            
            pdf.set_font("Helvetica", "", 9)
            campos_principal = [
                ("Archivo PDF", datos.get('archivo', 'N/A')),
                ("Número de Contrato", datos.get('numero_contrato', 'N/A')),
                ("Zona", datos.get('zona', 'N/A')),
            ]
            
            for label, valor in campos_principal:
                pdf.set_font("Helvetica", "B", 8.5)
                pdf.cell(40, 5, f"{label}:")
                pdf.set_font("Helvetica", "", 8.5)
                pdf.multi_cell(0, 5, str(valor)[:80], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            
            pdf.cell(0, 3, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            
            # CONTRATISTA GANADOR
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(0, 8, "CONTRATISTA GANADOR", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=0)
            
            pdf.set_font("Helvetica", "", 9)
            campos_ganador = [
                ("RUC", datos.get('ruc', 'N/A')),
                ("Razón Social", datos.get('contratista', 'N/A')),
                ("Representante Legal", datos.get('representante_legal', 'N/A')),
                ("DNI Representante", datos.get('dni_representante', 'N/A')),
            ]
            
            for label, valor in campos_ganador:
                pdf.set_font("Helvetica", "B", 8.5)
                pdf.cell(40, 5, f"{label}:")
                pdf.set_font("Helvetica", "", 8.5)
                pdf.multi_cell(0, 5, str(valor)[:80], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            
            pdf.cell(0, 3, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            
            # INFORMACIÓN OSCE DEL GANADOR
            if datos.get('ruc'):
                pdf.set_font("Helvetica", "B", 11)
                pdf.cell(0, 8, "INFORMACIÓN OSCE - EMPRESA GANADORA", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=0)
                
                pdf.set_font("Helvetica", "", 9)
                campos_osce = [
                    ("Razón Social OSCE", datos.get('osce_razon_social', 'N/A')),
                    ("Estado RNP", datos.get('osce_estado_rnp', 'N/A')),
                    ("Apto para contratar", "Sí" if datos.get('osce_es_apto') else "No"),
                    ("Teléfono(s)", ', '.join(datos.get('osce_telefonos', ['N/A']))),
                    ("Celular(es)", ', '.join(datos.get('osce_celulares', ['N/A']))),
                    ("Email(s)", ', '.join(datos.get('osce_emails', ['N/A']))),
                ]
                
                for label, valor in campos_osce:
                    pdf.set_font("Helvetica", "B", 8.5)
                    pdf.cell(40, 5, f"{label}:")
                    pdf.set_font("Helvetica", "", 8.5)
                    valor_str = str(valor)[:100]
                    pdf.multi_cell(0, 5, valor_str, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                pdf.cell(0, 3, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            
            # MIEMBROS DEL CONSORCIO
            miembros_texto = datos.get('miembros_consorcio', '').strip()
            if miembros_texto:
                pdf.set_font("Helvetica", "B", 11)
                pdf.cell(0, 8, "MIEMBROS DEL CONSORCIO", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=0)
                
                miembros_list = miembros_texto.split('\n')
                
                # Mostrar lista de miembros
                pdf.set_font("Helvetica", "", 9)
                for i, miembro_txt in enumerate(miembros_list, 1):
                    miembro_txt = miembro_txt.strip()
                    if miembro_txt:
                        # Extraer RUC del miembro
                        ruc_match = re.search(r'(\d{11})', miembro_txt)
                        if ruc_match:
                            ruc_miembro = ruc_match.group(1)
                            pdf.set_font("Helvetica", "B", 8.5)
                            pdf.cell(10, 6, f"{i}.")
                            pdf.set_font("Helvetica", "", 8.5)
                            pdf.multi_cell(0, 6, miembro_txt, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                        else:
                            pdf.cell(10, 6, f"{i}.")
                            pdf.multi_cell(0, 6, miembro_txt, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                # INFORMACIÓN OSCE DE MIEMBROS DEL CONSORCIO
                if 'osce_consorcio' in datos and datos['osce_consorcio']:
                    pdf.cell(0, 4, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                    pdf.set_font("Helvetica", "B", 11)
                    pdf.cell(0, 8, "INFORMACIÓN OSCE - MIEMBROS DEL CONSORCIO", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=0)
                    
                    for idx_miembro, osce_miembro in enumerate(datos['osce_consorcio'], 1):
                        pdf.set_font("Helvetica", "B", 9)
                        ruc_miembro = osce_miembro.get('ruc', 'N/A')
                        razon_miembro = osce_miembro.get('razon_social', 'N/A')
                        pdf.cell(0, 7, f"Miembro {idx_miembro}: RUC {ruc_miembro}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                        
                        pdf.set_font("Helvetica", "", 8)
                        info_miembro = [
                            ("Razón Social", razon_miembro),
                            ("Estado RNP", osce_miembro.get('estado_rnp', 'N/A')),
                            ("Apto para contratar", "Sí" if osce_miembro.get('es_apto') else "No"),
                            ("Teléfono", ', '.join(osce_miembro.get('telefonos_osce', ['N/A']))),
                            ("Celular", ', '.join(osce_miembro.get('celulares_osce', ['N/A']))),
                            ("Email", ', '.join(osce_miembro.get('emails', ['N/A']))),
                        ]
                        
                        for lbl, val in info_miembro:
                            pdf.set_font("Helvetica", "B", 7.5)
                            pdf.cell(15, 4, "")
                            pdf.cell(35, 4, f"{lbl}:")
                            pdf.set_font("Helvetica", "", 7.5)
                            val_str = str(val)[:90]
                            pdf.multi_cell(0, 4, val_str, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                        
                        pdf.cell(0, 2, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                pdf.cell(0, 5, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            
            # INFORMACIÓN DIGESA (si existe)
            if datos.get('digesa_registros'):
                pdf.set_font("Helvetica", "B", 11)
                pdf.cell(0, 8, "REGISTROS SANITARIOS (DIGESA)", new_x=XPos.LMARGIN, new_y=YPos.NEXT, border=0)
                
                pdf.set_font("Helvetica", "", 8)
                for reg in datos.get('digesa_registros', [])[:3]:  # Mostrar primeros 3
                    pdf.cell(0, 5, f"- Producto: {reg.get('producto', 'N/A')[:60]}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                    pdf.cell(5, 4, "")
                    pdf.multi_cell(0, 4, f"Estado: {reg.get('estado', 'N/A')} | Emision: {reg.get('fecha_emision', 'N/A')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                if len(datos.get('digesa_registros', [])) > 3:
                    pdf.cell(0, 4, f"... y {len(datos.get('digesa_registros', [])) - 3} más", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                
                pdf.cell(0, 3, "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=utiles.nombre_archivo('contratos_lote', 'pdf')
        )
    except Exception as e:
        print(f"Error generando PDF lote: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error generando PDF: {str(e)}'}), 500


# ─────────────────────────────────────────────
#  Consulta DIGESA por RUC (la implementacion vive en digesa.py)
# ─────────────────────────────────────────────

def consultar_digesa(ruc):
    return digesa_api.consultar(ruc)


# Nombre historico que usaban otros modulos.
consultar_digesa_por_ruc = consultar_digesa


@bp.route('/consultar_digesa', methods=['POST'])
def consultar_digesa_endpoint():
    body = request.json or {}
    ruc = (body.get('ruc') or '').strip()
    if not ruc:
        return jsonify({'error': 'Debes enviar un RUC.'}), 400
    try:
        return jsonify(consultar_digesa(ruc))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except requests.HTTPError as e:
        return jsonify({'error': f'Error HTTP DIGESA: {e}'}), 502
    except Exception as e:
        return jsonify({'error': f'No se pudo consultar DIGESA: {e}'}), 500

@bp.route('/consultar_osce', methods=['POST'])
def consultar_osce_endpoint():
    body = request.json or {}
    ruc = (body.get('ruc') or '').strip()
    if not ruc:
        return jsonify({'error': 'Debes enviar un RUC.'}), 400

    try:
        return jsonify(consultar_osce_por_ruc(ruc))
    except Exception as e:
        return jsonify({'error': str(e)}), 200
